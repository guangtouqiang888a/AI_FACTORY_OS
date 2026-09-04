# Entry 058A — READ-ONLY database provenance audit (do not mutate)
from __future__ import annotations

import json
import re
import sqlite3
from datetime import datetime
from pathlib import Path

ROOT = Path(r"D:/AI_FACTORY_OS")
DB = ROOT / "data" / "ai_factory.db"
RAW = ROOT / "data" / "raw" / "xianyu"
OUT = ROOT / "docs" / "07_AUDIT" / "database" / "_058a_audit_raw.json"

KEYWORDS = re.compile(
    r"sample|mock|demo|test|fixture|simulation|synthetic|fake|seed|generated|练习|模拟|测试",
    re.I,
)


def main() -> None:
    report: dict = {
        "db_path": str(DB.resolve()),
        "exists": DB.exists(),
        "size": DB.stat().st_size if DB.exists() else None,
        "mtime": datetime.fromtimestamp(DB.stat().st_mtime).isoformat() if DB.exists() else None,
        "tables": {},
        "products_sample": [],
        "products_keyword_counts": {},
        "products_price_dist": {},
        "titles_flagged": [],
        "urls_sample": [],
        "raw_files": [],
        "products_vs_raw": {},
    }

    conn = sqlite3.connect(f"file:{DB.as_posix()}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    tables = [
        r[0]
        for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        )
    ]
    for t in tables:
        count = conn.execute(f"SELECT COUNT(*) FROM [{t}]").fetchone()[0]
        cols = [r[1] for r in conn.execute(f"PRAGMA table_info([{t}])")]
        # try timestamp columns
        ts_cols = [c for c in cols if any(x in c.lower() for x in ("time", "date", "at", "created", "updated", "collect", "observ"))]
        earliest = latest = None
        for tc in ts_cols:
            try:
                row = conn.execute(
                    f"SELECT MIN([{tc}]), MAX([{tc}]) FROM [{t}] WHERE [{tc}] IS NOT NULL AND [{tc}] != ''"
                ).fetchone()
                if row and row[0]:
                    if earliest is None or str(row[0]) < str(earliest):
                        earliest = row[0]
                    if latest is None or str(row[1]) > str(latest):
                        latest = row[1]
            except sqlite3.Error:
                pass
        report["tables"][t] = {
            "row_count": count,
            "columns": cols,
            "earliest": earliest,
            "latest": latest,
            "timestamp_cols_checked": ts_cols,
        }

    # products deep dive
    n = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    report["products_exact_count"] = n
    for r in conn.execute(
        "SELECT id, platform_id, keyword, title, price, want_count, view_count, source_url, collect_date, seller, tags FROM products ORDER BY id"
    ):
        d = dict(r)
        report["products_sample"].append(d)
        kw = d.get("keyword") or ""
        report["products_keyword_counts"][kw] = report["products_keyword_counts"].get(kw, 0) + 1
        price = d.get("price")
        pk = str(price)
        report["products_price_dist"][pk] = report["products_price_dist"].get(pk, 0) + 1
        blob = " ".join(str(d.get(k) or "") for k in ("title", "keyword", "source_url", "seller", "tags"))
        if KEYWORDS.search(blob):
            report["titles_flagged"].append({"id": d["id"], "title": d["title"], "keyword": d["keyword"], "url": d["source_url"]})
        if d.get("source_url"):
            report["urls_sample"].append({"id": d["id"], "url": d["source_url"]})

    # unique urls
    urls = [u["url"] for u in report["urls_sample"]]
    report["unique_source_urls"] = sorted(set(urls))
    report["unique_source_url_count"] = len(set(urls))

    # 99.9 cohort
    rows99 = [dict(r) for r in conn.execute("SELECT * FROM products WHERE price = 99.9")]
    report["price_99_9_rows"] = rows99
    report["price_99_9_count"] = len(rows99)

    # scores sample
    if "scores" in tables:
        report["scores_sample"] = [dict(r) for r in conn.execute("SELECT * FROM scores LIMIT 5")]
        report["scores_count"] = conn.execute("SELECT COUNT(*) FROM scores").fetchone()[0]

    # publish_queue / market_events current
    for t in ("publish_queue", "publish_evidence", "market_events", "market_signals", "selection_results"):
        if t in tables:
            report["tables"][t]["preview"] = [dict(r) for r in conn.execute(f"SELECT * FROM [{t}] LIMIT 3")]

    conn.close()

    # raw files
    if RAW.exists():
        for p in sorted(RAW.rglob("*")):
            if p.is_file():
                info = {
                    "path": str(p.relative_to(ROOT)),
                    "size": p.stat().st_size,
                    "mtime": datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
                    "name": p.name,
                }
                if p.suffix.lower() in (".xlsx", ".xls"):
                    try:
                        import openpyxl

                        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                        sheets = []
                        for name in wb.sheetnames:
                            ws = wb[name]
                            rows = list(ws.iter_rows(values_only=True))
                            header = rows[0] if rows else None
                            sheets.append(
                                {
                                    "sheet": name,
                                    "nrows": len(rows),
                                    "ncols": len(header) if header else 0,
                                    "header": [str(x) for x in (header or [])],
                                    "sample_rows": [
                                        [str(c) if c is not None else None for c in row]
                                        for row in rows[1:4]
                                    ],
                                }
                            )
                        info["workbook"] = {"sheets": sheets, "creator": getattr(wb.properties, "creator", None)}
                        wb.close()
                    except Exception as e:
                        info["workbook_error"] = str(e)
                report["raw_files"].append(info)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT)
    print("products_exact_count", report["products_exact_count"])
    print("tables", {k: v["row_count"] for k, v in report["tables"].items()})
    print("unique_urls", report["unique_source_url_count"], report["unique_source_urls"][:10])
    print("price_dist", report["products_price_dist"])
    print("flagged", len(report["titles_flagged"]))
    print("raw_files", len(report["raw_files"]))


if __name__ == "__main__":
    main()
