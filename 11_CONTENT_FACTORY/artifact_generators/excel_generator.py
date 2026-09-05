# 11_CONTENT_FACTORY/artifact_generators/excel_generator.py — 真实 XLSX 生成

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import Any


def generate_excel(title: str, sheets: list[dict], output_path: Path) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        return {"status": "error", "file_path": "", "error": f"openpyxl not installed: {exc}"}

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def.get("name", "Sheet")[:31])
        columns = sheet_def.get("columns", [])
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        sample_rows = sheet_def.get("sample_rows", 3)
        for i in range(1, sample_rows + 1):
            row = []
            for col in columns:
                if col in ("金额", "数值", "合计"):
                    row.append(i * 100)
                elif col == "日期":
                    row.append(f"2026-07-{i:02d}")
                else:
                    row.append(f"示例{i}")
            ws.append(row)

        formulas = sheet_def.get("formulas", [])
        if "SUM" in formulas and len(columns) >= 2:
            sum_col = 2 if len(columns) > 1 else 1
            ws.cell(row=sample_rows + 3, column=1, value="合计")
            col_letter = chr(64 + sum_col) if sum_col <= 26 else "B"
            ws.cell(row=sample_rows + 3, column=sum_col, value=f"=SUM({col_letter}2:{col_letter}{sample_rows + 1})")

        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = max(12, len(str(col_name)) + 4)

    if not wb.sheetnames:
        wb.create_sheet("数据")

    wb.properties.title = title
    wb.save(str(output_path))
    return {"status": "ok", "file_path": str(output_path)}


def generate_project_plan_gantt_excel(title: str, output_path: Path) -> dict[str, Any]:
    """
    Deterministic sellable Excel: 使用说明 + 任务明细 + 项目摘要 + 甘特图.
    No LLM. AI cost = 0.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule
    except ImportError as exc:
        return {"status": "error", "file_path": "", "error": f"openpyxl not installed: {exc}"}

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    gantt_fill = PatternFill("solid", fgColor="5B9BD5")
    guide_fill = PatternFill("solid", fgColor="E2EFDA")

    # --- 使用说明 ---
    ws_guide = wb.active
    ws_guide.title = "使用说明"
    guide_lines = [
        ["小微团队项目计划 + 任务进度 + 甘特图 Excel 模板", ""],
        ["【填写区说明】黄色单元格 = 请用户填写；灰色/蓝色 = 自动计算或展示", ""],
        ["1. 打开「任务明细」工作表，从第2行开始填写任务。", ""],
        ["2. 必填：任务名称、负责人、开始日期、截止日期、状态、完成进度(0-1)。", ""],
        ["3. 「项目摘要」会自动统计任务数、完成数、平均进度。", ""],
        ["4. 「甘特图」按开始/截止日期自动显示时间条（与任务明细日期一致）。", ""],
        ["5. 可增删任务行：在任务明细第2–31行范围内填写即可。", ""],
        ["6. 日期请使用 YYYY-MM-DD 格式（如 2026-09-05）。", ""],
        ["7. 完成进度填写 0 到 1 之间的小数，例如 0.5 表示 50%。", ""],
        ["8. 本模板面向小微团队/项目负责人，无需复杂项目管理软件。", ""],
        ["【注意】本商品为可编辑数字模板；售出后不退不换。", ""],
    ]
    for r, row in enumerate(guide_lines, 1):
        ws_guide.cell(row=r, column=1, value=row[0])
        if r == 1:
            ws_guide.cell(row=r, column=1).font = Font(bold=True, size=14, color="1F4E79")
        else:
            ws_guide.cell(row=r, column=1).fill = guide_fill
    ws_guide.column_dimensions["A"].width = 88

    # --- 任务明细 ---
    ws_task = wb.create_sheet("任务明细")
    headers = ["任务ID", "任务名称", "负责人", "开始日期", "截止日期", "状态", "完成进度", "工期(天)", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws_task.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ("T01", "明确项目目标与范围", "张三", date(2026, 9, 1), date(2026, 9, 3), "已完成", 1.0, "示例，可改"),
        ("T02", "拆解任务并分配负责人", "李四", date(2026, 9, 3), date(2026, 9, 6), "进行中", 0.6, "示例，可改"),
        ("T03", "跟踪进度并更新甘特图", "王五", date(2026, 9, 5), date(2026, 9, 12), "未开始", 0.0, "示例，可改"),
        ("T04", "周会复盘与风险记录", "张三", date(2026, 9, 8), date(2026, 9, 8), "未开始", 0.0, "示例，可改"),
        ("T05", "交付验收与归档", "李四", date(2026, 9, 10), date(2026, 9, 15), "未开始", 0.0, "示例，可改"),
    ]
    for i, (tid, name, owner, start, end, status, prog, note) in enumerate(samples, 2):
        values = [tid, name, owner, start, end, status, prog, None, note]
        for c, v in enumerate(values, 1):
            cell = ws_task.cell(row=i, column=c, value=v)
            cell.border = thin
            if c in (2, 3, 4, 5, 6, 7, 9):
                cell.fill = input_fill
            if c in (4, 5):
                cell.number_format = "YYYY-MM-DD"
            if c == 7:
                cell.number_format = "0%"
        # 工期公式
        ws_task.cell(row=i, column=8, value=f'=IF(OR(D{i}="",E{i}=""),"",E{i}-D{i}+1)')
        ws_task.cell(row=i, column=8).border = thin

    # empty input rows 7-21 for user
    for i in range(7, 22):
        ws_task.cell(row=i, column=1, value=f"=IF(B{i}=\"\",\"\",\"T\"&TEXT({i}-1,\"00\"))")
        for c in range(1, 10):
            cell = ws_task.cell(row=i, column=c)
            cell.border = thin
            if c in (2, 3, 4, 5, 6, 7, 9):
                cell.fill = input_fill
            if c in (4, 5):
                cell.number_format = "YYYY-MM-DD"
            if c == 7:
                cell.number_format = "0%"
            if c == 8:
                cell.value = f'=IF(OR(D{i}="",E{i}=""),"",E{i}-D{i}+1)'
        # fix T01 formula for empty rows - simpler task id
        ws_task.cell(row=i, column=1, value="")
        ws_task.cell(row=i, column=1).fill = input_fill
        ws_task.cell(row=i, column=1).border = thin

    # re-fill sample task ids without formulas for clarity
    for i, (tid, *_rest) in enumerate(samples, 2):
        ws_task.cell(row=i, column=1, value=tid)
        ws_task.cell(row=i, column=1).fill = input_fill
        ws_task.cell(row=i, column=1).border = thin

    widths = [10, 28, 12, 12, 12, 10, 10, 10, 16]
    for i, w in enumerate(widths, 1):
        ws_task.column_dimensions[get_column_letter(i)].width = w

    # --- 项目摘要 ---
    ws_sum = wb.create_sheet("项目摘要")
    ws_sum["A1"] = "项目总体进度摘要"
    ws_sum["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_sum["A3"] = "项目名称（请填写）"
    ws_sum["B3"] = "我的小微团队项目"
    ws_sum["B3"].fill = input_fill
    ws_sum["A4"] = "项目负责人（请填写）"
    ws_sum["B4"] = "张三"
    ws_sum["B4"].fill = input_fill
    ws_sum["A6"] = "任务总数"
    ws_sum["B6"] = '=COUNTA(任务明细!B2:B21)'
    ws_sum["A7"] = "已完成任务数"
    ws_sum["B7"] = '=COUNTIF(任务明细!F2:F21,"已完成")'
    ws_sum["A8"] = "进行中任务数"
    ws_sum["B8"] = '=COUNTIF(任务明细!F2:F21,"进行中")'
    ws_sum["A9"] = "未开始任务数"
    ws_sum["B9"] = '=COUNTIF(任务明细!F2:F21,"未开始")'
    ws_sum["A10"] = "平均完成进度"
    ws_sum["B10"] = '=IFERROR(AVERAGE(任务明细!G2:G21),0)'
    ws_sum["B10"].number_format = "0%"
    ws_sum["A12"] = "最早开始日期"
    ws_sum["B12"] = '=IFERROR(MIN(任务明细!D2:D21),"")'
    ws_sum["B12"].number_format = "YYYY-MM-DD"
    ws_sum["A13"] = "最晚截止日期"
    ws_sum["B13"] = '=IFERROR(MAX(任务明细!E2:E21),"")'
    ws_sum["B13"].number_format = "YYYY-MM-DD"
    ws_sum["A15"] = "使用提示"
    ws_sum["A16"] = "先在「任务明细」填写任务；本页自动汇总；再看「甘特图」时间条。"
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 28

    # --- 甘特图 ---
    ws_g = wb.create_sheet("甘特图")
    ws_g["A1"] = "任务名称"
    ws_g["B1"] = "开始"
    ws_g["C1"] = "截止"
    for c in range(1, 4):
        ws_g.cell(row=1, column=c).fill = header_fill
        ws_g.cell(row=1, column=c).font = header_font
        ws_g.cell(row=1, column=c).border = thin

    # timeline: 14 days from project min start hypothesis 2026-09-01
    base = date(2026, 9, 1)
    day_count = 16
    for d in range(day_count):
        col = 4 + d
        cell = ws_g.cell(row=1, column=col, value=base + timedelta(days=d))
        cell.number_format = "M/D"
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", textRotation=90)
        ws_g.column_dimensions[get_column_letter(col)].width = 3.5

    ws_g.column_dimensions["A"].width = 28
    ws_g.column_dimensions["B"].width = 12
    ws_g.column_dimensions["C"].width = 12

    for i in range(2, 22):
        # link to task sheet
        ws_g.cell(row=i, column=1, value=f'=IF(任务明细!B{i}="","",任务明细!B{i})')
        ws_g.cell(row=i, column=2, value=f'=IF(任务明细!D{i}="","",任务明细!D{i})')
        ws_g.cell(row=i, column=2).number_format = "YYYY-MM-DD"
        ws_g.cell(row=i, column=3, value=f'=IF(任务明细!E{i}="","",任务明细!E{i})')
        ws_g.cell(row=i, column=3).number_format = "YYYY-MM-DD"
        for c in range(1, 4):
            ws_g.cell(row=i, column=c).border = thin
        for d in range(day_count):
            col = 4 + d
            # day date in header row 1
            # show bar if day between start and end inclusive
            # D$header is date in row 1
            col_letter = get_column_letter(col)
            formula = (
                f'=IF(OR($A{i}="",$B{i}="",$C{i}=""),"",'
                f'IF(AND({col_letter}$1>=$B{i},{col_letter}$1<=$C{i}),"█",""))'
            )
            cell = ws_g.cell(row=i, column=col, value=formula)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
            cell.font = Font(color="1F4E79")

    ws_g["A24"] = "说明：蓝色「█」表示该任务在对应日期处于计划区间内；与「任务明细」开始/截止日期保持一致。"
    ws_g["A24"].font = Font(italic=True, color="666666")

    wb.properties.title = title
    wb.properties.creator = "AI_FACTORY_OS Content Factory (deterministic)"
    wb.save(str(output_path))
    return {"status": "ok", "file_path": str(output_path), "generator": "project_plan_gantt"}
